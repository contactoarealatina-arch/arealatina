"""Generación de planillas Excel con openpyxl.

Todas las funciones devuelven bytes listos para servir como descarga.
"""
from io import BytesIO

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Colores de la marca aplicados a la planilla.
NARANJA = 'FF5722'
NEGRO = '1A1A1A'
GRIS = 'F2F2F2'

FUENTE_TITULO = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
RELLENO_TITULO = PatternFill('solid', fgColor=NARANJA)
BORDE_FINO = Border(*[Side(style='thin', color='DDDDDD')] * 4)


def _nueva_hoja(titulo, encabezados):
    libro = Workbook()
    hoja = libro.active
    hoja.title = titulo[:31]  # Excel no admite nombres de hoja más largos.

    hoja.append(encabezados)
    for celda in hoja[1]:
        celda.font = FUENTE_TITULO
        celda.fill = RELLENO_TITULO
        celda.alignment = Alignment(horizontal='center', vertical='center')
    hoja.freeze_panes = 'A2'
    return libro, hoja


def _ajustar(hoja, minimo=10, maximo=45):
    """Ancho de columna según el contenido más largo."""
    for columna in hoja.columns:
        largo = max((len(str(c.value)) for c in columna if c.value is not None), default=0)
        letra = get_column_letter(columna[0].column)
        hoja.column_dimensions[letra].width = max(minimo, min(largo + 3, maximo))


def _guardar(libro):
    buffer = BytesIO()
    libro.save(buffer)
    return buffer.getvalue()


def _formato_filas(hoja, primera=2):
    for fila in hoja.iter_rows(min_row=primera):
        for celda in fila:
            celda.border = BORDE_FINO


# ---------------------------------------------------------------------------
# Alumnos
# ---------------------------------------------------------------------------
def exportar_alumnos(alumnos):
    libro, hoja = _nueva_hoja('Alumnos', [
        'Nombre completo', 'RUT', 'Estado', 'Teléfono', 'Email',
        'Fecha de ingreso', 'Clases', 'Plan actual', 'Vence', 'Estado de pago',
        'Contacto de emergencia', 'Teléfono emergencia',
    ])

    for alumno in alumnos:
        suscripcion = alumno.suscripcion_vigente
        hoja.append([
            alumno.nombre_completo,
            alumno.rut,
            alumno.get_estado_display(),
            alumno.telefono,
            alumno.email,
            alumno.fecha_ingreso,
            ', '.join(i.clase.get_nombre_display() for i in alumno.inscripciones.all()),
            suscripcion.plan.nombre if suscripcion else '',
            suscripcion.fecha_vencimiento if suscripcion else '',
            alumno.estado_pago_display,
            alumno.contacto_emergencia,
            alumno.telefono_emergencia,
        ])

    for fila in hoja.iter_rows(min_row=2, min_col=6, max_col=6):
        for celda in fila:
            celda.number_format = 'DD/MM/YYYY'
    for fila in hoja.iter_rows(min_row=2, min_col=9, max_col=9):
        for celda in fila:
            celda.number_format = 'DD/MM/YYYY'

    _formato_filas(hoja)
    _ajustar(hoja)
    return _guardar(libro)


# ---------------------------------------------------------------------------
# Pagos
# ---------------------------------------------------------------------------
def exportar_pagos(pagos, titulo='Pagos'):
    libro, hoja = _nueva_hoja(titulo, [
        'Fecha', 'Alumno', 'RUT', 'Concepto', 'Monto (CLP)',
        'Método', 'Comprobante', 'Estado', 'Registrado por',
    ])

    total = 0
    for pago in pagos:
        hoja.append([
            pago.fecha_pago,
            pago.alumno.nombre_completo,
            pago.alumno.rut,
            pago.concepto_display,
            pago.monto_clp,
            pago.get_metodo_display(),
            pago.numero_comprobante,
            pago.get_estado_display(),
            pago.registrado_por.get_full_name() if pago.registrado_por else '',
        ])
        if pago.estado == 'PAGADO':
            total += pago.monto_clp

    ultima = hoja.max_row + 1
    hoja.cell(row=ultima, column=4, value='Total cobrado').font = Font(bold=True)
    celda_total = hoja.cell(row=ultima, column=5, value=total)
    celda_total.font = Font(bold=True)

    for fila in hoja.iter_rows(min_row=2, min_col=1, max_col=1):
        for celda in fila:
            celda.number_format = 'DD/MM/YYYY'
    for fila in hoja.iter_rows(min_row=2, min_col=5, max_col=5):
        for celda in fila:
            celda.number_format = '#,##0'

    _formato_filas(hoja)
    _ajustar(hoja)
    return _guardar(libro)


# ---------------------------------------------------------------------------
# Asistencia
# ---------------------------------------------------------------------------
def exportar_asistencia(registros):
    libro, hoja = _nueva_hoja('Asistencia', [
        'Fecha', 'Clase', 'Nivel', 'Alumno', 'RUT', 'Estado', 'Observación',
    ])

    for registro in registros:
        hoja.append([
            registro.fecha,
            registro.clase.get_nombre_display(),
            registro.clase.get_nivel_display(),
            registro.alumno.nombre_completo,
            registro.alumno.rut,
            registro.get_estado_display(),
            registro.observacion,
        ])

    for fila in hoja.iter_rows(min_row=2, min_col=1, max_col=1):
        for celda in fila:
            celda.number_format = 'DD/MM/YYYY'

    _formato_filas(hoja)
    _ajustar(hoja)
    return _guardar(libro)


# ---------------------------------------------------------------------------
# Próximos vencimientos
# ---------------------------------------------------------------------------
def exportar_vencimientos(suscripciones):
    libro, hoja = _nueva_hoja('Por vencer', [
        'Alumno', 'RUT', 'Teléfono', 'Plan', 'Inicio', 'Vence', 'Días restantes',
    ])

    for sus in suscripciones:
        hoja.append([
            sus.alumno.nombre_completo,
            sus.alumno.rut,
            sus.alumno.telefono,
            sus.plan.nombre,
            sus.fecha_inicio,
            sus.fecha_vencimiento,
            sus.dias_restantes,
        ])

    for columna in (5, 6):
        for fila in hoja.iter_rows(min_row=2, min_col=columna, max_col=columna):
            for celda in fila:
                celda.number_format = 'DD/MM/YYYY'

    _formato_filas(hoja)
    _ajustar(hoja)
    return _guardar(libro)


# ---------------------------------------------------------------------------
# Ingresos mensuales, con gráfico incrustado
# ---------------------------------------------------------------------------
def exportar_ingresos(serie, titulo='Ingresos mensuales'):
    """serie = [{'etiqueta': 'ago 26', 'total': 350000}, ...]"""
    libro, hoja = _nueva_hoja('Ingresos', ['Mes', 'Ingresos (CLP)'])

    for punto in serie:
        hoja.append([punto['etiqueta'], punto['total']])

    for fila in hoja.iter_rows(min_row=2, min_col=2, max_col=2):
        for celda in fila:
            celda.number_format = '#,##0'

    _formato_filas(hoja)
    _ajustar(hoja)

    if serie:
        grafico = BarChart()
        grafico.title = titulo
        grafico.y_axis.title = 'CLP'
        grafico.x_axis.title = 'Mes'
        grafico.height = 8
        grafico.width = 18

        datos = Reference(hoja, min_col=2, min_row=1, max_row=len(serie) + 1)
        categorias = Reference(hoja, min_col=1, min_row=2, max_row=len(serie) + 1)
        grafico.add_data(datos, titles_from_data=True)
        grafico.set_categories(categorias)
        hoja.add_chart(grafico, 'D2')

    return _guardar(libro)


# ---------------------------------------------------------------------------
# Historial de la profesora
# ---------------------------------------------------------------------------
def exportar_historial_profesora(filas, titulo_mes):
    libro, hoja = _nueva_hoja('Historial', [
        'Fecha', 'Clase', 'Nivel', 'Profesor/a',
        'Presentes', 'Ausentes', 'Justificados', 'Total', '% asistencia',
    ])

    for fila in filas:
        clase = fila['clase']
        hoja.append([
            fila['fecha'],
            clase.get_nombre_display(),
            clase.get_nivel_display(),
            clase.profesora.get_full_name() if clase.profesora else 'Sin asignar',
            fila['presentes'],
            fila['ausentes'],
            fila['justificados'],
            fila['total'],
            fila['porcentaje'] / 100,
        ])

    for fila in hoja.iter_rows(min_row=2, min_col=1, max_col=1):
        for celda in fila:
            celda.number_format = 'DD/MM/YYYY'
    for fila in hoja.iter_rows(min_row=2, min_col=9, max_col=9):
        for celda in fila:
            celda.number_format = '0%'

    if filas:
        total_marcas = sum(f['total'] for f in filas)
        total_presentes = sum(f['presentes'] for f in filas)
        ultima = hoja.max_row + 1
        hoja.cell(row=ultima, column=1, value=f'Resumen {titulo_mes}').font = Font(bold=True)
        hoja.cell(row=ultima, column=5, value=total_presentes).font = Font(bold=True)
        hoja.cell(row=ultima, column=8, value=total_marcas).font = Font(bold=True)
        celda = hoja.cell(row=ultima, column=9,
                          value=(total_presentes / total_marcas) if total_marcas else 0)
        celda.number_format = '0%'
        celda.font = Font(bold=True)

    _formato_filas(hoja)
    _ajustar(hoja)
    return _guardar(libro)
